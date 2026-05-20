"""
build_tcap_master.py
Reads all TCAP district-level files (2015–2025), normalizes columns,
extracts MSCS (system=792) and TN State average, saves tcap_master.xlsx.

Metric note:
  2015–2021  → pct_on_mastered  ("On Track or Mastered", old TNReady scale)
  2022–2025  → pct_met_exceeded ("Met or Exceeded Expectations", new benchmark)
  These two metrics are NOT directly comparable. The series is shown in full with
  a discontinuity marker at 2022.
"""

import pandas as pd
import numpy as np
from pathlib import Path
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.utils import get_column_letter

DATA_DIR = Path(__file__).parent
OUT      = DATA_DIR / "tcap_master.xlsx"

TEAL    = "FF0F6E56"
LIGHT   = "FFE8F5F0"
WHITE   = "FFFFFFFF"
AMBER   = "FFFFF3CD"   # highlight for benchmark-change years
RED_BG  = "FFFCE8E6"   # highlight for missing data

SUBJECTS_MAP = {
    "rla": "ELA", "ela": "ELA", "english language arts": "ELA",
    "math": "Math", "mathematics": "Math",
    "science": "Science",
    "social studies": "Social Studies",
}
KEEP_SUBJECTS = {"ELA", "Math", "Science", "Social Studies"}

MSCS_ID     = 792
SPECIAL_MIN = 900   # systems >= 900 are state programs — exclude from TN avg

# school_year label for each data-year integer
SCHOOL_YEAR = {
    2015: "2014-15",
    2016: "2015-16",
    2017: "2016-17",
    2018: "2017-18",
    2019: "2018-19",
    2020: "2019-20",
    2021: "2020-21",
    2022: "2021-22",
    2023: "2022-23",
    2024: "2023-24",
    2025: "2024-25",
}

# ── file catalogue ─────────────────────────────────────────────────────────────
FILES = {
    2015: ("data_2015_district_base.xlsx",                     "xlsx"),
    # 2016 = EOC only (Algebra I / English I, grades 9-12) — no TCAP 3-8; excluded
    2017: ("data_2017_district_base.csv",                      "csv"),
    2018: ("data_2018_district_base.csv",                      "csv"),
    2019: ("district_assessment_file_suppressed.csv",          "csv"),
    # 2020 → COVID assessment waiver — no data
    2021: ("district_assessment_file_suppressed_upd422.csv",   "csv"),
    2022: ("district_assessment_file_suppressed_upd32323.xlsx","xlsx"),
    2023: ("district_assessment_file_suppressed_2023.xlsx",    "xlsx"),
    2024: ("district_assessment_file_suppressed_2024.xlsx",    "xlsx"),
    2025: ("district_assessment_file_suppressed_2025.xlsx",    "xlsx"),
}

METRIC_ERA = {yr: ("On Track/Mastered (TNReady)" if yr <= 2021 else "Met/Exceeded (new benchmark)")
              for yr in FILES}

# ── loaders ───────────────────────────────────────────────────────────────────

def load_raw(year, fname, ftype):
    path = DATA_DIR / fname
    if ftype == "csv":
        df = pd.read_csv(path, low_memory=False)
    else:
        df = pd.read_excel(path)
    df.columns = [c.strip() for c in df.columns]
    return df


def normalize(year, df):
    """Return DataFrame with standardized columns; print column diagnostics."""
    cols = {c.lower(): c for c in df.columns}

    print(f"\n{'─'*60}")
    print(f"  FILE: {FILES[year][0]}")
    print(f"  SCHOOL YEAR: {SCHOOL_YEAR.get(year, str(year))}")
    print(f"  RAW COLUMNS ({len(df.columns)}):")
    for i, c in enumerate(df.columns, 1):
        print(f"    {i:2d}. {c}")

    # ── system id ─────────────────────────────────────────────────────────────
    sys_col  = next((cols[k] for k in ("system","district number","district_number") if k in cols), None)
    name_col = next((cols[k] for k in ("system_name","district name","district_name") if k in cols), None)

    # ── subgroup ──────────────────────────────────────────────────────────────
    if year >= 2022:
        sub_col = next((cols[k] for k in ("student_group","student group") if k in cols), None)
    else:
        sub_col = next((cols[k] for k in ("subgroup",) if k in cols), None)

    # ── subject ───────────────────────────────────────────────────────────────
    subj_col = next((cols[k] for k in ("subject",) if k in cols), None)

    # ── grade ─────────────────────────────────────────────────────────────────
    grade_col = next((cols[k] for k in ("grade","grade_band","grade band") if k in cols), None)

    # ── proficiency metric ─────────────────────────────────────────────────────
    if year == 2015:
        pct_col = next((cols[k] for k in ("pct_prof_adv","% proficient/advanced") if k in cols), None)
        metric_name = "pct_prof_adv (Proficient+Advanced)"
    elif year <= 2021:
        pct_col = next((cols[k] for k in ("pct_on_mastered",) if k in cols), None)
        metric_name = "pct_on_mastered (On Track or Mastered)"
    else:
        pct_col = next((cols[k] for k in ("pct_met_exceeded",) if k in cols), None)
        metric_name = "pct_met_exceeded (Met or Exceeded Expectations)"

    # ── test type (TNReady vs Alt) ─────────────────────────────────────────────
    test_col = next((cols[k] for k in ("test",) if k in cols), None)

    print(f"\n  MAPPED COLUMNS:")
    print(f"    system_id   → {sys_col}")
    print(f"    system_name → {name_col}")
    print(f"    subgroup    → {sub_col}")
    print(f"    subject     → {subj_col}")
    print(f"    grade       → {grade_col}")
    print(f"    proficiency → {pct_col}  [{metric_name}]")
    print(f"    test_type   → {test_col}")

    if not all([sys_col, sub_col, subj_col, pct_col]):
        print(f"  ⚠ SKIPPING {year} — missing required column(s)")
        return pd.DataFrame()

    n = len(df)
    subject_raw = df[subj_col].astype(str).str.strip()
    out = pd.DataFrame({
        "year":        [year] * n,
        "school_year": [SCHOOL_YEAR.get(year, str(year))] * n,
        "system":      pd.to_numeric(df[sys_col], errors="coerce").values,
        "system_name": df[name_col].astype(str).str.strip().values if name_col else [""] * n,
        "subgroup":    df[sub_col].astype(str).str.strip().values,
        "subject_raw": subject_raw.values,
        "subject":     subject_raw.str.lower().map(SUBJECTS_MAP).values,
        "grade":       df[grade_col].astype(str).str.strip().values if grade_col else ["All Grades"] * n,
        "pct":         pd.to_numeric(df[pct_col], errors="coerce").values,
        "test":        df[test_col].astype(str).str.strip().values if test_col else ["TNReady"] * n,
        "metric_era":  [METRIC_ERA[year]] * n,
    })
    return out


def filter_rows(df):
    """Keep All Students, All Grades (3-8 aggregate), KEEP_SUBJECTS, TNReady/ACH only."""
    if df.empty:
        return df
    df = df[df["subgroup"].str.lower().isin(["all students","all","all students (non-tested enrolled)"])]
    df = df[df["subject"].isin(KEEP_SUBJECTS)]
    df = df[df["grade"].str.lower().isin(["all grades","all","3-8","grades 3-8",""])]
    df = df[df["system"].notna() & (df["system"] > 0)]
    if "test" in df.columns:
        main = df[df["test"].str.strip().str.lower().isin(["tnready","ach"])]
        if not main.empty:
            df = main
    return df


# ── sheet builders ─────────────────────────────────────────────────────────────

def build_master():
    frames = []
    for year, (fname, ftype) in FILES.items():
        raw    = load_raw(year, fname, ftype)
        norm   = normalize(year, raw)
        filt   = filter_rows(norm)
        print(f"\n  ROWS AFTER FILTER: {len(filt)}")
        frames.append(filt)
    master = pd.concat([f for f in frames if not f.empty], ignore_index=True)
    master = master[["year","school_year","system","system_name","subject","grade",
                     "subgroup","pct","metric_era"]].copy()
    master = master.dropna(subset=["year"])
    master["year"] = master["year"].astype(int)
    return master


def trend_sheet(master):
    """TREND sheet: one row per (school_year, entity), columns = subjects."""
    rows = []
    for yr in sorted(master["year"].unique()):
        sy = SCHOOL_YEAR.get(yr, str(yr))
        era = METRIC_ERA[yr]
        yr_data = master[master["year"] == yr]
        # MSCS
        mscs = yr_data[yr_data["system"] == MSCS_ID]
        for entity_label, subset in [("Memphis-Shelby County Schools", mscs),
                                      ("TN State Avg (all districts)", yr_data[(yr_data["system"] < SPECIAL_MIN) & (yr_data["system"] != MSCS_ID)])]:
            row = {"school_year": sy, "year": yr, "entity": entity_label, "metric_era": era}
            for subj in sorted(KEEP_SUBJECTS):
                s = subset[subset["subject"] == subj]["pct"]
                row[subj] = round(s.mean(), 1) if not s.empty and s.notna().any() else None
            rows.append(row)
    df = pd.DataFrame(rows, columns=["school_year","year","entity","metric_era",
                                     "ELA","Math","Science","Social Studies"])
    return df


def subject_sheet(master, subj):
    """One row per school_year: MSCS %, TN Avg %, Gap, missing flags."""
    rows = []
    for yr in sorted(master["year"].unique()):
        sy  = SCHOOL_YEAR.get(yr, str(yr))
        era = METRIC_ERA[yr]
        sub = master[(master["year"] == yr) & (master["subject"] == subj)]
        mscs_v = sub[sub["system"] == MSCS_ID]["pct"].mean()
        tn_v   = sub[(sub["system"] < SPECIAL_MIN) & (sub["system"] != MSCS_ID)]["pct"].mean()
        mscs_flag = "⚠ suppressed/missing" if pd.isna(mscs_v) else ""
        tn_flag   = "⚠ suppressed/missing" if pd.isna(tn_v)   else ""
        rows.append({
            "School Year":       sy,
            "Year":              yr,
            "MSCS %":            round(mscs_v, 1) if pd.notna(mscs_v) else None,
            "TN Avg %":          round(tn_v, 1)   if pd.notna(tn_v)   else None,
            "Gap (MSCS – TN)":   round(mscs_v - tn_v, 1) if (pd.notna(mscs_v) and pd.notna(tn_v)) else None,
            "MSCS Flag":         mscs_flag,
            "TN Flag":           tn_flag,
            "Metric Era":        era,
        })
    return pd.DataFrame(rows)


def all_districts_sheet(master):
    """ALL_DISTRICTS: every district × every year × every subject."""
    out = master[master["system"] < SPECIAL_MIN].copy()
    out["school_year"] = out["year"].map(SCHOOL_YEAR)
    pivot = out.pivot_table(index=["year","school_year","system","system_name"],
                            columns="subject", values="pct", aggfunc="mean")
    pivot = pivot.reindex(columns=sorted(KEEP_SUBJECTS))
    pivot = pivot.reset_index()
    pivot.columns.name = None
    return pivot.sort_values(["year","system"])


def metadata_sheet():
    rows = [
        {"School Year":"2014-15","Year":2015,"File":"data_2015_district_base.xlsx",
         "Metric":"pct_prof_adv","Era":"Old TCAP (Proficient/Advanced)","Notes":"RLA→ELA; no Social Studies"},
        {"School Year":"2015-16","Year":2016,"File":"data_2016_suppressed_district_base.xlsx",
         "Metric":"N/A","Era":"EOC only","Notes":"Algebra I / English I grades 9-12 only — EXCLUDED (no TCAP 3-8)"},
        {"School Year":"2016-17","Year":2017,"File":"data_2017_district_base.csv",
         "Metric":"pct_on_mastered","Era":"TNReady (On Track or Mastered)","Notes":"No Social Studies"},
        {"School Year":"2017-18","Year":2018,"File":"data_2018_district_base.csv",
         "Metric":"pct_on_mastered","Era":"TNReady","Notes":""},
        {"School Year":"2018-19","Year":2019,"File":"district_assessment_file_suppressed.csv",
         "Metric":"pct_on_mastered","Era":"TNReady","Notes":"MSCS Science suppressed"},
        {"School Year":"2019-20","Year":2020,"File":"N/A",
         "Metric":"N/A","Era":"N/A","Notes":"COVID assessment waiver — no data"},
        {"School Year":"2020-21","Year":2021,"File":"district_assessment_file_suppressed_upd422.csv",
         "Metric":"pct_on_mastered","Era":"TNReady","Notes":"Low participation due to COVID"},
        {"School Year":"2021-22","Year":2022,"File":"district_assessment_file_suppressed_upd32323.xlsx",
         "Metric":"pct_met_exceeded","Era":"NEW BENCHMARK — Met/Exceeded Expectations","Notes":"⚠ BENCHMARK CHANGE — not directly comparable to 2015-2021"},
        {"School Year":"2022-23","Year":2023,"File":"district_assessment_file_suppressed_2023.xlsx",
         "Metric":"pct_met_exceeded","Era":"Met/Exceeded","Notes":""},
        {"School Year":"2023-24","Year":2024,"File":"district_assessment_file_suppressed_2024.xlsx",
         "Metric":"pct_met_exceeded","Era":"Met/Exceeded","Notes":""},
        {"School Year":"2024-25","Year":2025,"File":"district_assessment_file_suppressed_2025.xlsx",
         "Metric":"pct_met_exceeded","Era":"Met/Exceeded","Notes":""},
    ]
    meta = pd.DataFrame(rows)
    meta.loc[len(meta)] = {
        "School Year":"","Year":None,"File":"Built by","Metric":"Mahendra Gudipadu",
        "Era":"Data Analyst, Peer Power Foundation",
        "Notes":"Source: TDOE Data Downloads — tn.gov/education/districts/federal-programs-and-oversight/data/data-downloads.html"
    }
    return meta


# ── Excel formatting ───────────────────────────────────────────────────────────

def style_sheet(ws, benchmark_change_rows=None):
    teal_fill  = PatternFill("solid", fgColor=TEAL)
    light_fill = PatternFill("solid", fgColor=LIGHT)
    white_fill = PatternFill("solid", fgColor=WHITE)
    amber_fill = PatternFill("solid", fgColor="FFFFF3CD")
    bold_white = Font(bold=True, color="FFFFFFFF", size=11)

    for cell in ws[1]:
        cell.fill = teal_fill
        cell.font = bold_white
        cell.alignment = Alignment(horizontal="center", vertical="center")

    for row_idx, row in enumerate(ws.iter_rows(min_row=2), start=2):
        if benchmark_change_rows and row_idx in benchmark_change_rows:
            fill = amber_fill
        else:
            fill = light_fill if row_idx % 2 == 0 else white_fill
        for cell in row:
            cell.fill   = fill
            cell.alignment = Alignment(horizontal="center")

    for col in ws.columns:
        max_len = max((len(str(c.value)) if c.value is not None else 0) for c in col)
        ws.column_dimensions[get_column_letter(col[0].column)].width = min(max_len + 4, 44)

    ws.freeze_panes = "A2"


def pct_format(ws, pct_cols_set, df):
    for i, col in enumerate(df.columns, 1):
        if col in pct_cols_set:
            letter = get_column_letter(i)
            for cell in ws[letter]:
                if cell.row == 1:
                    continue
                if cell.value is not None:
                    cell.number_format = "0.0"


def write_sheet(writer, df, sheet_name, pct_cols=None, benchmark_rows=None):
    df.to_excel(writer, sheet_name=sheet_name, index=False)
    ws = writer.sheets[sheet_name]
    style_sheet(ws, benchmark_change_rows=benchmark_rows)
    if pct_cols:
        pct_format(ws, pct_cols, df)


# ── summary printer ────────────────────────────────────────────────────────────

def print_summary(master):
    print("\n" + "═" * 100)
    print("  MSCS vs TN STATE AVERAGE — ALL YEARS ALL SUBJECTS")
    print("  Gap = MSCS minus TN Avg. Negative = MSCS below state avg.")
    print("  ⚠ = suppressed / missing data  |  ★ = benchmark changed (2022+, not directly comparable)")
    print("═" * 100)
    hdr = (f"{'Year':>5}  {'School Yr':>9}  "
           f"{'ELA MSCS':>9} {'ELA TN':>7} {'Gap':>6}  "
           f"{'Math MSCS':>9} {'Math TN':>7} {'Gap':>6}  "
           f"{'Sci MSCS':>9} {'Sci TN':>7} {'Gap':>6}  "
           f"{'SS MSCS':>8} {'SS TN':>7} {'Gap':>6}  "
           f"{'Notes'}")
    print(hdr)
    print("─" * 100)

    all_years = sorted(master["year"].unique())
    # insert gaps for 2016 and 2020
    display_years = sorted(set(all_years) | {2016, 2020})

    for yr in display_years:
        sy    = SCHOOL_YEAR.get(yr, str(yr))
        flags = []

        if yr == 2016:
            print(f"  {yr:>4}  {sy:>9}  {'EOC ONLY — no TCAP 3-8 data':^80}")
            continue
        if yr == 2020:
            print(f"  {yr:>4}  {sy:>9}  {'COVID WAIVER — no assessment data':^80}")
            continue
        if yr == 2022:
            flags.append("★ benchmark change")

        yr_data = master[master["year"] == yr]

        def pct(subset, subj):
            v = subset[subset["subject"] == subj]["pct"].mean()
            return v

        mscs_m = yr_data["system"] == MSCS_ID
        tn_m   = (yr_data["system"] < SPECIAL_MIN) & (yr_data["system"] != MSCS_ID)

        parts = []
        for subj in ["ELA", "Math", "Science", "Social Studies"]:
            mv = pct(yr_data[mscs_m], subj)
            tv = pct(yr_data[tn_m],   subj)
            ms = f"{mv:6.1f}" if pd.notna(mv) else "   N/A"
            ts = f"{tv:6.1f}" if pd.notna(tv) else "   N/A"
            if pd.isna(mv):
                flags.append(f"⚠ {subj} MSCS suppressed")
            if pd.isna(tv):
                flags.append(f"⚠ {subj} TN suppressed")
            if pd.notna(mv) and pd.notna(tv):
                gap_v = mv - tv
                gap_s = f"{gap_v:+5.1f}"
            else:
                gap_s = "   N/A"
            parts.append((ms, ts, gap_s))

        ela, math, sci, ss = parts
        flag_str = "  " + " | ".join(flags) if flags else ""
        print(f"  {yr:>4}  {sy:>9}  "
              f"{ela[0]:>9} {ela[1]:>7} {ela[2]:>6}  "
              f"{math[0]:>9} {math[1]:>7} {math[2]:>6}  "
              f"{sci[0]:>9} {sci[1]:>7} {sci[2]:>6}  "
              f"{ss[0]:>8} {ss[1]:>7} {ss[2]:>6}"
              f"{flag_str}")

    print("═" * 100)
    print("  METRIC ERAS:")
    print("  2014-15 to 2018-19, 2020-21  →  On Track or Mastered (pct_on_mastered) / TNReady")
    print("  2021-22 to 2024-25           →  Met or Exceeded Expectations (pct_met_exceeded) ★ NEW BENCHMARK")
    print("  2019-20                      →  No data (COVID assessment waiver)")
    print("  2015-16                      →  No TCAP 3-8 data (EOC only file)")
    print("═" * 100)


# ── main ──────────────────────────────────────────────────────────────────────

def main():
    print("=" * 60)
    print("Building TCAP Master — all years (2015–2025)")
    print("=" * 60)

    master = build_master()

    print(f"\n{'═'*60}")
    print(f"  MASTER DATASET SUMMARY")
    print(f"  Total rows: {len(master):,}")
    dist_by_yr = master.groupby("year")["system"].nunique().rename("districts")
    print(f"\n  Districts per year:")
    for yr, cnt in dist_by_yr.items():
        print(f"    {yr} ({SCHOOL_YEAR.get(yr,'?')}): {cnt} districts")

    mscs_check = master[master["system"] == MSCS_ID].groupby(["year","subject"])["pct"].first().unstack()
    print(f"\n  MSCS (system=792) proficiency values by year × subject:")
    print(mscs_check.to_string())

    # Build sheets
    trend_df  = trend_sheet(master)
    ela_df    = subject_sheet(master, "ELA")
    math_df   = subject_sheet(master, "Math")
    sci_df    = subject_sheet(master, "Science")
    ss_df     = subject_sheet(master, "Social Studies")
    all_df    = all_districts_sheet(master)
    meta_df   = metadata_sheet()

    subj_pct  = {"ELA", "Math", "Science", "Social Studies"}
    flat_pct  = {"MSCS %", "TN Avg %", "Gap (MSCS – TN)"}

    # Identify benchmark-change row indices for amber highlight (year 2022 rows)
    def bm_rows(df, year_col="year"):
        return {i + 2 for i, yr in enumerate(df[year_col]) if yr == 2022}

    with pd.ExcelWriter(OUT, engine="openpyxl") as writer:
        write_sheet(writer, trend_df, "TREND",          pct_cols=subj_pct,  benchmark_rows=bm_rows(trend_df))
        write_sheet(writer, ela_df,   "ELA",            pct_cols=flat_pct,  benchmark_rows=bm_rows(ela_df, "Year"))
        write_sheet(writer, math_df,  "MATH",           pct_cols=flat_pct,  benchmark_rows=bm_rows(math_df, "Year"))
        write_sheet(writer, sci_df,   "SCIENCE",        pct_cols=flat_pct,  benchmark_rows=bm_rows(sci_df, "Year"))
        write_sheet(writer, ss_df,    "SOCIAL_STUDIES", pct_cols=flat_pct,  benchmark_rows=bm_rows(ss_df, "Year"))
        write_sheet(writer, all_df,   "ALL_DISTRICTS",  pct_cols=subj_pct)
        write_sheet(writer, meta_df,  "METADATA")

    print(f"\n✅  Saved: {OUT}")

    print_summary(master)


if __name__ == "__main__":
    main()
